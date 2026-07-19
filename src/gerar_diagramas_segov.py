import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ARQUIVO_SAIDA = "Estrutura_Completa_8_Processos_Visio.xlsx"

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Paleta de estilo padrão governamental (SEGOV)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_data = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Engenharia de processos com base no decreto consolidado para o órgão
processos_completos = {
    "01. Tombamento": [
        ("T1", "Recebimento do Bem (Nota Fiscal, Doação ou Cessão)", "T2", "Início", "Start", "Almoxarifado"),
        ("T2", "Conferência Física do Item vs. Documento Fiscal", "T3", "Ok", "Decision", "Almoxarifado"),
        ("T3", "Definição da Conta Contábil, Vida Útil e Valor Residual", "T4", "Conforme", "Process", "Patrimônio Setorial"),
        ("T4", "Geração do Número Patrimonial e Impressão de Plaqueta/QR Code", "T5", "", "Process", "Patrimônio Setorial"),
        ("T5", "Fixação Física da Plaqueta de Identificação no Bem", "T6", "", "Process", "Almoxarifado"),
        ("T6", "Vinculação do Centro de Custo, Localização e Responsável no Sistema", "T7", "", "Process", "Patrimônio Setorial"),
        ("T7", "Emissão, Assinatura e Arquivamento do Termo de Guarda e Responsabilidade", "T8", "", "Process", "Área Demandante / Usuário"),
        ("T8", "Incorporação ao Ativo Imobilizado e Liberação para Uso", "", "Fim", "End", "Patrimônio Setorial"),
    ],
    "02. Inventário": [
        ("I1", "Publicação de Portaria e Instituição da Comissão de Inventário", "I2", "Início", "Start", "Diretoria Executiva / Dirigente"),
        ("I2", "Emissão de Relatório Teórico e Cronograma de Varredura", "I3", "", "Process", "Patrimônio Setorial"),
        ("I3", "Vistoria em Campo e Leitura Física dos Códigos de Barra/QR Code", "I4", "", "Process", "Comissão de Inventário"),
        ("I4", "Avaliação do Estado de Conservação (Bom, Ocioso, Inservível, etc.)", "I5", "", "Process", "Comissão de Inventário"),
        ("I5", "Confronto e Conciliação: Dados de Campo vs. Base Sistêmica", "I6", "", "Decision", "Patrimônio Setorial"),
        ("I6", "Instauração de Sindicância ou Regularização de Bens Achados", "I7", "Divergente", "Process", "Diretoria Executiva / Dirigente"),
        ("I7", "Elaboração, Homologação e Assinatura do Relatório Final do Inventário", "I8", "Ok", "Process", "Comissão de Inventário"),
        ("I8", "Atualização Cadastral e Renovação Geral dos Termos de Responsabilidade", "", "Fim", "End", "Patrimônio Setorial"),
    ],
    "03. Cessão de Uso": [
        ("C1", "Recebimento de Solicitação Formal de Órgão Cessionário Externo", "C2", "Início", "Start", "Diretoria Executiva / Dirigente"),
        ("C2", "Análise Técnica de Ociosidade e Viabilidade Operacional do Bem", "C3", "", "Process", "Patrimônio Setorial"),
        ("C3", "Análise Jurídica e Confecção da Minuta do Termo de Cessão de Uso", "C4", "", "Process", "Assessoria Jurídica"),
        ("C4", "Aprovação do Parecer Jurídico e Assinatura do Termo pelas Autoridades", "C5", "Aprovado", "Decision", "Diretoria Executiva / Dirigente"),
        ("C5", "Alteração do Status do Bem para 'Cedido' e Vinculação da Carga Externa", "C6", "", "Process", "Patrimônio Setorial"),
        ("C6", "Entrega Física do Ativo Mediante Assinatura do Termo de Recebimento", "", "Fim", "End", "Almoxarifado"),
    ],
    "04. Doação de Bens": [
        ("D1", "Abertura de Processo com Laudo de Bens Inservíveis ou Desnecessários", "D2", "Início", "Start", "Patrimônio Setorial"),
        ("D2", "Autorização da Alta Administração para Doação com Fim Social", "D3", "Autorizado", "Decision", "Diretoria Executiva / Dirigente"),
        ("D3", "Chamamento Público ou Análise de Regularidade Jurídica do Donatário", "D4", "Ok", "Process", "Assessoria Jurídica"),
        ("D4", "Assinatura do Termo de Doação Definitiva pelas Partes", "D5", "", "Process", "Diretoria Executiva / Dirigente"),
        ("D5", "Processamento da Baixa Patrimonial por Doação no Sistema", "D6", "", "Process", "Patrimônio Setorial"),
        ("D6", "Emissão da Nota Fiscal de Baixa e Retirada dos Bens pelo Donatário", "", "Fim", "End", "Almoxarifado"),
    ],
    "05. Desfazimento e Baixa": [
        ("B1", "Identificação de Sinistro (Roubo/Furto), Destruição ou Obsolescência", "B2", "Início", "Start", "Patrimônio Setorial"),
        ("B2", "Emissão de Laudo de Inservibilidade ou Anexação de Boletim de Ocorrência", "B3", "", "Process", "Comissão de Descarte / Técnico"),
        ("B3", "Julgamento, Homologação e Autorização da Baixa pela Diretoria", "B4", "Aprovado", "Decision", "Diretoria Executiva / Dirigente"),
        ("B4", "Descaracterização Física Completa do Bem (Remoção de Logos e Marcas)", "B5", "", "Process", "Almoxarifado"),
        ("B5", "Descarte Ecológico Homologado (Geração de CDF) ou Inclusão em Leilão", "B6", "", "Process", "Almoxarifado"),
        ("B6", "Conciliação e Lançamento da Baixa Contábil Definitiva no ERP", "", "Fim", "End", "Contabilidade Geral"),
    ],
    "06. Bem Particular": [
        ("P1", "Solicitação de Entrada e Uso de Equipamento Pessoal a Trabalho", "P2", "Início", "Start", "Área Demandante / Usuário"),
        ("P2", "Análise de Segurança da Informação, Redes e Requisitos Técnicos", "P3", "Aprovado", "Decision", "Tecnologia da Informação (TI)"),
        ("P3", "Assinatura do Termo de Cautela, Autorização de Uso e Isenção de Danos", "P4", "", "Process", "Área Demandante / Usuário"),
        ("P4", "Fixação de Etiqueta de Identificação Provisória de Item Particular", "P5", "", "Process", "Patrimônio Setorial"),
        ("P5", "Uso Autorizado do Bem nas Dependências do Órgão Público", "P6", "", "Process", "Área Demandante / Usuário"),
        ("P6", "Solicitação de Desvinculação ou Desligamento do Colaborador", "P7", "", "Process", "Área Demandante / Usuário"),
        ("P7", "Baixa na Cautela, Remoção da Identificação e Liberação na Portaria", "", "Fim", "End", "Patrimônio Setorial"),
    ],
    "07. Transferência": [
        ("F1", "Abertura de Solicitação de Remanejamento Interno de Ativos", "F2", "Início", "Start", "Gestor do Setor de Origem"),
        ("F2", "Liberação Sistêmica e Deslocamento Físico do Bem para Nova Unidade", "F3", "", "Process", "Almoxarifado"),
        ("F3", "Conferência Física do Estado e Características do Bem no Destino", "F4", "", "Process", "Gestor do Setor de Destino"),
        ("F4", "Aceite Eletrônico da Carga Patrimonial no Sistema", "F5", "Confirmado", "Decision", "Gestor do Setor de Destino"),
        ("F5", "Atualização Automática do Centro de Custo no Banco de Dados", "F6", "", "Process", "Patrimônio Setorial"),
        ("F6", "Geração e Assinatura Digital do Novo Termo de Responsabilidade", "", "Fim", "End", "Patrimônio Setorial"),
    ],
    "08. Reavaliação e Depreciação": [
        ("R1", "Fechamento Mensal dos Movimentos e Histórico de Bens", "R2", "Início", "Start", "Patrimônio Setorial"),
        ("R2", "Cálculo Automatizado das Quotas de Depreciação Linear Acumulada", "R3", "", "Process", "Patrimônio Setorial"),
        ("R3", "Aplicação Periódica do Teste de Recuperabilidade de Ativos (Impairment)", "R4", "", "Process", "Patrimônio Setorial"),
        ("R4", "Necessidade de Ajuste ao Valor Justo de Mercado Encontrada?", "R5", "Sim", "Decision", "Patrimônio Setorial"),
        ("R5", "Contratação e Emissão de Laudo Técnico por Perito Especializado", "R6", "", "Process", "Contabilidade Geral"),
        ("R6", "Integração Contábil dos Ajustes e Saldos no Balanço Patrimonial do Órgão", "", "Fim", "End", "Contabilidade Geral"),
    ],
}

headers = [
    "ID do Passo",
    "Descrição do Passo",
    "Próximo Passo",
    "Texto do Conector",
    "Tipo de Forma (Shape)",
    "Dono / Raia (Swimlane)",
]


def ajustar_colunas(ws):
    for coluna in ws.columns:
        largura = 12
        letra = get_column_letter(coluna[0].column)
        for cell in coluna:
            if cell.value:
                largura = max(largura, min(len(str(cell.value)) + 2, 70))
        ws.column_dimensions[letra].width = largura


for sheet_name, rows in processos_completos.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True
    ws["A1"] = f"Mapeamento Oficial: {sheet_name[4:]}"
    ws["A1"].font = font_title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    for col_num, title in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num, value=title)
        cell.font = font_header
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(rows, start=5):
        fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font_data
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in [1, 3, 5] else "left",
                vertical="top",
                wrap_text=True,
            )

    ws.freeze_panes = "A5"
    ajustar_colunas(ws)

wb.save(ARQUIVO_SAIDA)
print(f"Arquivo '{ARQUIVO_SAIDA}' gerado com absoluto sucesso!")
